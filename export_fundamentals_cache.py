#!/usr/bin/env python3
"""
从财务 Excel 文件中提取股票的净利同比增长和总市值数据，生成 JSON 缓存。

注意：这是 Q1 季度数据，非最新季度！

输出: stock_fundamentals_cache.json
用法: python3 export_fundamentals_cache.py
"""
import os
import json
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE, "final_rankings_with_details_Q1.xlsx")
CACHE_FILE = os.path.join(BASE, "stock_fundamentals_cache.json")

def export():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Excel 不存在: {EXCEL_FILE}")
        return

    print("读取 Excel...")
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)  # data_only=True 读值不读公式
    ws = wb.active
    rows = ws.max_row
    print(f"共 {rows} 行，提取中...")

    cache = {}
    COL_CODE = 0       # A 代码
    COL_JLZZ = 5       # F 净利同比增长 (Q1)
    COL_MCAP = 13      # N 总市值 (Q1)

    for r in range(2, rows + 1):
        code_raw = ws.cell(row=r, column=COL_CODE + 1).value
        if not code_raw:
            continue
        code = str(code_raw).zfill(6)

        # 净利同比增长 (Q1)
        jl_raw = ws.cell(row=r, column=COL_JLZZ + 1).value
        jl_growth = None
        if jl_raw and jl_raw != '-':
            try:
                jl_growth = float(str(jl_raw).replace('%', ''))
            except (ValueError, TypeError):
                pass

        # 总市值 (Q1) (格式如 "68亿")
        mcap_raw = ws.cell(row=r, column=COL_MCAP + 1).value
        mcap_yi = None
        if mcap_raw and mcap_raw != '-':
            try:
                s = str(mcap_raw).lower()
                if '亿' in s:
                    mcap_yi = float(s.replace('亿', '').strip())
                elif '万' in s:
                    mcap_yi = float(s.replace('万', '').strip()) / 10000
            except (ValueError, TypeError):
                pass

        if jl_growth is not None or mcap_yi is not None:
            cache[code] = {}
            if jl_growth is not None:
                cache[code]['jl_growth_pct'] = round(jl_growth, 2)
            if mcap_yi is not None:
                cache[code]['mcap_yi'] = mcap_yi

    wb.close()

    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    # 统计
    jl_vals = [v['jl_growth_pct'] for v in cache.values() if 'jl_growth_pct' in v]
    mc_vals = [v['mcap_yi'] for v in cache.values() if 'mcap_yi' in v]
    pos = sum(1 for v in jl_vals if v > 0) if jl_vals else 0

    print(f"💾 已保存 {len(cache)} 只")
    print(f"   净利同比正增长: {pos} 只")
    if mc_vals:
        mc_sorted = sorted(mc_vals)
        print(f"   市值中位数: {mc_sorted[len(mc_sorted)//2]:.1f}亿")

if __name__ == "__main__":
    export()
